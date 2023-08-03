import openpyxl
import os


class LoadData:
    def __init__(self):
        print(os.getcwd())
        self.path = os.getcwd()
        self.path = os.path.join(self.path, "Courses Data.xlsx")
        print(self.path)
        self.wb = openpyxl.load_workbook(self.path)
        self.ws = self.wb.active

    def getCoursesList(self):
        data = []

        for row in self.ws.iter_rows(min_row=2, max_col=5, max_row=100):
            if (
                (row[1].value == None)
                or (row[0].value == None)
                or (row[2].value == None)
            ):
                break
            course = row[1].value + " - " + row[0].value + " (" + row[2].value + ")"
            # print(course)
            data.append(course)
        return data

    # def getCertificateType(self):
    #     data = {}
    #     for row in self.ws.iter_rows(min_row=2, max_col=6, max_row=100):
    #         if (row[1].value == None) or (row[5].value == None):
    #             break
    #         if row[5].value not in data:
    #             data[row[5].value] = []
    #         data[row[5].value].append(row[1].value)
    #     # print(data)
    #     return data

    def getCertificateType(self, courseName):
        for row in self.ws.iter_rows(min_row=2, max_col=6, max_row=100):
            if (row[1].value == None) or (row[5].value == None):
                break
            if row[1].value == courseName:
                return row[5].value

    def coursesWithNoExpireDate(self):
        data = []
        for row in self.ws.iter_rows(min_row=2, max_col=6, max_row=100):
            if (row[2].value == None) or (row[3].value == None):
                break
            if row[3].value == False:
                data.append(row[2].value)
        # print(data)
        return data

    def getRulesCode(self, courseName):
        for row in self.ws.iter_rows(min_row=2, max_col=6, max_row=100):
            if (row[1].value == None) or (row[4].value == None):
                break
            if courseName == row[1].value:
                return row[4].value

if __name__ == "__main__":
    loader = LoadData()
    print(loader.getCoursesList())
    print(loader.getCertificateType("التعامل مع الحوادث الكيميائية"))
    print(loader.coursesWithNoExpireDate())
    print(loader.getRulesCode("التعامل مع الحوادث الكيميائية"))