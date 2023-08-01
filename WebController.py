import openpyxl
import os
import shutil


class WebController:
    def __init__(self, workingDir):
        self.workingDir = workingDir
        self.excelPath = workingDir + "/WebSheet.xlsx"
        try:
            self.wb = openpyxl.load_workbook(self.excelPath)

        except:
            shutil.copy("Website Template.xlsx", self.excelPath)
            self.wb = openpyxl.load_workbook(self.excelPath)

        self.ws = self.wb.active

    def getWorkingRow(self):
        workingRow = 2
        while True:
            if self.ws["C" + str(workingRow)].value == None:
                break
            workingRow += 1
        return workingRow

    def addDateToExcel(self, data):
        certType_En = "Mandatory"
        certType_Ar = "الحتمية"
        working_row = self.getWorkingRow()

        self.ws["A" + str(working_row)] = certType_Ar
        self.ws["B" + str(working_row)] = certType_En
        self.ws["C" + str(working_row)] = data["courseCode"] + str(data["CertNo"])
        self.ws["D" + str(working_row)] = data["RegNo"]
        self.ws["E" + str(working_row)] = data["Name_Ar"]
        self.ws["F" + str(working_row)] = data["Name_En"]
        self.ws["G" + str(working_row)] = data["Place_of_Birth_Ar"]
        self.ws["H" + str(working_row)] = data["Place_of_Birth_En"]
        self.ws["I" + str(working_row)] = data["Date_of_Birth"]
        self.ws["I" + str(working_row)].number_format = "dd/mm/yyyy"
        self.ws["J" + str(working_row)] = data["FromWhere_Ar"]
        self.ws["K" + str(working_row)] = data["FromWhere_En"]
        self.ws["L" + str(working_row)] = data["From"]
        self.ws["L" + str(working_row)].number_format = "dd/mm/yyyy"
        self.ws["M" + str(working_row)] = data["To"]
        self.ws["M" + str(working_row)].number_format = "dd/mm/yyyy"
        self.ws["N" + str(working_row)] = data["Course_Ar"]
        self.ws["O" + str(working_row)] = data["Course_En"]
        self.ws["P" + str(working_row)] = data["Issue_date"]
        self.ws["P" + str(working_row)].number_format = "dd/mm/yyyy"
        self.ws["Q" + str(working_row)] = data["Expire_date"]
        self.ws["Q" + str(working_row)].number_format = "dd/mm/yyyy"

        self.wb.save(self.excelPath)


# data = {
#     "courseCode": "XNFF",
#     "CertNo": 123,
#     "Name_En": "Omar Ahmed Mahmoud Ahmed Mahmoud Shalaby".upper(),
#     "Name_Ar": "عمر أحمد محمود أحمد محمود شلبي",
#     "Place_of_Birth_En": "Alex".upper(),
#     "Place_of_Birth_Ar": "الإسكندرية",
#     "Date_of_Birth": "01/01/1999",
#     "FromWhere_En": "Personal",
#     "FromWhere_Ar": "شخصي",
#     "Course_En": "Fire Fighting",
#     "Course_Ar": "الإطفاء",
#     "From": "01/01/2020",
#     "To": "02/01/2020",
#     "Rules": "A - VII / 1-1",
#     "Expire_date": "01/01/2025",
#     "Issue_date": "04/01/2020",
#     "RegNo": "H00123456",
# }

# import LoadData as ld

# loader = ld.LoadData()
# courseslist = loader.getCoursesList()
# # print(courseslist)
# Courses_En = [x.split(" - ")[0] for x in courseslist]
# Courses_Ar = [x.split(" - ")[1].split(" (")[0] for x in courseslist]
# Courses_Code = [x.split(" - ")[1].split(" (")[-1].split(")")[0] for x in courseslist]
# # print(Courses_Code)

# for i in range(len(Courses_En)):
#     data["courseCode"] = Courses_Code[i]
#     data["Course_En"] = Courses_En[i]
#     data["Course_Ar"] = Courses_Ar[i]
#     data["Rules"] = loader.getRulesCode(Courses_En[i])
#     writer = WebController("D:/Projects/Python/MTUC Project/Results")
#     writer.addDateToExcel(data)
#     print(f"{i} Done")

# websheet = WebController("D:\Projects\Python\MTUC Project\Results")
# websheet.addDateToExcel(data)
